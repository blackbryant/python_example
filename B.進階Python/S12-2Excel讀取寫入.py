#Excel檔案存取
#
import openpyxl

#寫入xlsx
workbook =openpyxl.Workbook()
sheet = workbook.worksheets[0]
sheet['A1'] = 'hello'
listtitle =['姓名','電話']
sheet.append(listtitle)
workbook.save("test2.xlsx")

#讀取Excel
workbook=openpyxl.load_workbook("test2.xlsx")
sheet=workbook.worksheets[0]
for i in range(1,sheet.max_row+1):
    for j in range(1,sheet.max_column+1):
        print(sheet.cell(row=i,column=j).value,end="    ")
    print()
sheet['A3']='Perry'
workbook.save('test2.xlsx')

